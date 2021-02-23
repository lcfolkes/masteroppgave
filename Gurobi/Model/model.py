# imports
import os
import sys
import gurobipy as gp
from gurobipy import GRB, Model
import numpy as np
from itertools import product
from Gurobi.Model import config_reader
import pandas as pd
from openpyxl import load_workbook
print(sys.path)#os.getcwd())

# TODO: add error message when file does not exist (attribute error)
#input_file = "4-10-1-1.txt"
directory = "../tests/4nodes/"
#filepath = os.path.join(directory, input_file)
files = []
for filename in os.listdir(directory):
    files.append(os.path.join(directory, filename))

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]
        Returns: None
        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
            add_header = False
        except FileNotFoundError:
            # file does not exist yet, we will create it
            add_header = True
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, header=add_header, **to_excel_kwargs)

        # save the workbook
        writer.save()

def run_test_instance(filepath):
    cp = config_reader.ConfigReader(filepath)

    # print("SCENARIOS: ", SCENARIOS)
    NODES = np.arange(1, cp.numNodes+1) # N, set of nodes
    # print("NODES: ", NODES)
    CHARGING_NODES = np.arange(cp.numPNodes + 1, cp.numPNodes+cp.numCNodes + 1) # N^C, set of charging nodes
    # print("CHARGING NODES: ", CHARGING_NODES)
    PARKING_NODES = np.arange(1, cp.numPNodes+1) # N^P, set of parking nodes
    # print("PARKING NODES: ", PARKING_NODES)
    # NPS = list(np.arange(NUM_SURPLUS_NODES)) # N^(P+), set of surplus nodes
    PARKING_NEED_CHARGING_NODES = cp.carsInNeedCTranslate # N^(PC), set of parking nodes with cars in need of charging
    # print("PARKING_NEED_CHARGING_NODES: ", PARKING_NEED_CHARGING_NODES)
    CARS = np.arange(1, cp.numCars+1) # C, set of cars potentially subject to relocation
    # print("CARS: ", CARS)
    CARMOVES = np.arange(1, cp.numCarMovesP + cp.numCarMovesC+1) # R, set of car-moves
    CARMOVES_CAR = config_reader.create_dict_of_indices(CARS, cp.carMoveCars)  # R_c, set of car-moves for car c
    PARKING_MOVES = np.arange(1, cp.numCarMovesP+1)  # set of PARKING-moves, R_i^PD and R_i^PO
    # print(PARKING_MOVES)
    CHARGING_MOVES = np.arange(cp.numCarMovesP+1, cp.numCarMoves+1)  # set of charging-moves, trenger  R_i^CD and R_i^CO
    EMPLOYEES = np.arange(1, cp.numEmployees+1) # K, set of service employees
    TASKS = np.arange(1, cp.numTasks+1) # M, set of possible abstract tasks

    # DECLARE PARAMETERS
    #SCENARIO_PROBABILITY = dict(zip(SCENARIOS, cp.scenarioProb)) # P_s, Probability of scenario s
    # print("SCENARIO_PROBABILITY: ", SCENARIO_PROBABILITY)
    PROFIT_RENTAL = cp.profitRental # C^(RC), Profit per unit of time of renting out cars
    COST_RELOCATION = cp.costRelocation # C^R, Cost per unit of time of relocation activities
    COST_DEVIATION = cp.costDeviation    # C^D, Cost per car of not reaching ideal state
    INITIAL_AVAILABLE_CARS = dict(zip(PARKING_NODES, cp.carsAvailableInNode)) # S^(In)_i, Initial number of available cars (cars not in need of charging)
    IDEAL_STATE = dict(zip(PARKING_NODES, cp.idealStateP))   # S_i^(P) , Ideal state in parking node i
    INITIAL_NEED_CHARGING = dict(zip(PARKING_NEED_CHARGING_NODES, cp.carsInNeedNodes)) # S_i^C, Initial number of cars that require charging at parking node i
    INITIAL_AVAILABLE_CAPACITY = dict(zip(CHARGING_NODES, cp.chargingSlotsAvailable)) # N_i^(CS), Initial available capacity of charging node i
    TASKS_FIRST_STAGE = np.arange(1, cp.numTasksFirstStage+1) # F, number of tasks in first stage, check that F is smaller or equal to len(M)
    TASKS_SECOND_STAGE = np.arange(cp.numTasksFirstStage+1, cp.numTasks+1)
    PARKING_MOVES_ORIGINATING_IN_NODE, PARKING_MOVES_ENDING_IN_NODE, CHARGING_MOVES_ORIGINATING_IN_NODE, CHARGING_MOVES_ENDING_IN_NODE \
        = config_reader.create_car_moves_origin_destination(PARKING_NODES, CHARGING_NODES, cp.carMoveOrigin, cp.carMoveDestination)
    # R_i^(PO), R_i^(PD), R_i^(CO), R_i^(CD)
    CARMOVE_ORIGIN = dict(zip(CARMOVES, cp.carMoveOrigin))  # o(r), origin node of car move r
    CARMOVE_DESTINATION = dict(zip(CARMOVES, cp.carMoveDestination))  # d(r), destination node of car move r
    CARMOVE_START_TIME = dict(zip(CARMOVES, cp.carMoveStartTime))
    RELOCATION_TIME = dict(zip(CARMOVES, cp.carMoveHandlingTime))  # T_r^H, time needed to perform car-move r
    START_TIME_EMPLOYEE = dict(zip(EMPLOYEES, cp.travelTimeToOrigin))  # T_k^(SO), earliest start time of service employee k
    EMPLOYEE_START_LOCATION = dict(zip(EMPLOYEES, cp.startNodeEmployee))  # o(k), location of employee k at time T^(SO)_k
    TRAVEL_TIME = cp.travelTimeBike  # T_(ij), travel times between node i and j
    PLANNING_PERIOD = cp.timeLimit  # T^bar, planning period
    BIGM = dict(zip(CARMOVES, cp.bigM))
    scenarios = np.arange(1, cp.numScenarios+1)

    CUSTOMER_REQUESTS={}
    CUSTOMER_DELIVERIES={}

    #Model type
    #0: stochastic
    #1: deterministic
    #3: Expectation of the expected value problem
    def create_model(modelType, input_model=None):
        if modelType==0:
            print("\nCreating stochastic Model")
        elif modelType==1:
            print("\nCreating deterministic Model")
        elif modelType==2:
            if input_model is not None:
                print("\nCreating EEV Model")
            else:
                print("\nNeed input_model argument")
                return

        # INITIALIZE NODE SETS
        if modelType == 0 or modelType == 3:
            SCENARIOS = np.arange(1, cp.numScenarios+1) # S, set of scenarios
            SCENARIO_PROBABILITY = 1 / cp.numScenarios

        else:
            SCENARIOS = np.arange(1, 2)
            SCENARIO_PROBABILITY = 1

        if modelType==0 or modelType == 3:
            CUSTOMER_REQUESTS = config_reader.read_2d_array_to_dict(cp.demandP) # R_(is), number of customer requests in second stage in parking node i in scenario s in second stage
            CUSTOMER_DELIVERIES = config_reader.read_2d_array_to_dict(cp.deliveriesP) # D_(is), number of vehicles delivered by customers in node i and scenario s in second stage
        else:
            deterministic_customer_request = {}
            deterministic_customer_deliveries = {}
            for i in range(len(PARKING_NODES)):
                deterministic_customer_request[(i+1,1)] = 1#round(sum(cp.demandP[i])/len(cp.demandP[i]),0)
                deterministic_customer_deliveries[(i+1,1)] = 1#round(sum(cp.deliveriesP[i])/len(cp.deliveriesP[i]),0)
            CUSTOMER_REQUESTS = deterministic_customer_request
            CUSTOMER_DELIVERIES = deterministic_customer_deliveries

        # Create a new Model
        m = gp.Model("mip1")

        # Create variables
        x = m.addVars(product(EMPLOYEES, CARMOVES, TASKS, SCENARIOS), vtype=GRB.BINARY, name="x") # x_krms, 1 if service employee k performs car-move r as task number m in scenario s, 0 otherwise
        if modelType == 3:
            for v in input_model.getVars():
                # x[krms], m < tasksfirsstage
                # t[kms], m < tasksfirststage
                if v.varName[0] == 'x':
                    var_indices = list(map(int, (v.varName[2:-1].split(','))))
                    if var_indices[2] < len(TASKS_FIRST_STAGE) + 1:
                        for s in SCENARIOS:
                            var_indices[-1] = s
                            x[tuple(var_indices)].lb = v.x
                            x[tuple(var_indices)].ub = v.x

        y = m.addVars(PARKING_NODES, lb=0, vtype=GRB.INTEGER, name="y") # y_i, Number of cars in node i by the beginning of the second stage
        z = m.addVars(product(PARKING_NODES, SCENARIOS), lb=0, vtype=GRB.INTEGER, name="z") # z_is, number of customer requests served in second stage in node i in scenario s
        w = m.addVars(product(PARKING_NODES, SCENARIOS), lb=0, vtype=GRB.INTEGER, name="w") # w_is, number of cars short of ideal state in node i in scenario s
        t = m.addVars(product(EMPLOYEES, TASKS, SCENARIOS), vtype=GRB.CONTINUOUS, lb=0, name="t")
        # print(x)
        # print(y)
        # print(z)
        # print(w)

        ### OBJECTIVE FUNCTION ###
        # profit from customer requests
        profit_customer_requests = gp.quicksum((PROFIT_RENTAL)*z[(i,s)] for i in PARKING_NODES for s in SCENARIOS)

        # costs from relocation activities
        costs_relocation = gp.quicksum(COST_RELOCATION*RELOCATION_TIME[r]*x[(k,r,m,s)]
                                        for m in TASKS
                                        for r in CARMOVES
                                        for k in EMPLOYEES
                                        for s in SCENARIOS)

        # ideal state deviation cost
        cost_deviation_ideal_state = gp.quicksum(COST_DEVIATION*w[(i, s)] for i in PARKING_NODES for s in SCENARIOS)

        total_profit = SCENARIO_PROBABILITY*(profit_customer_requests - costs_relocation - cost_deviation_ideal_state)

        m.setObjective(total_profit, GRB.MAXIMIZE)

        ### CONSTRAINTS ###

        ## Relocation of rental vehicles

        # (2) Ensure each car is moved at most once
        m.addConstrs(
            (gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in CARMOVES_CAR[c] for m in TASKS) <= 1 for s in SCENARIOS for c in CARS), name="c2"
        )
        #print("c2")

        # (3) Make sure each task of each employee in every scenario consists of at most one car-move
        m.addConstrs(
            (gp.quicksum(x[(k,r,m,s)] for r in CARMOVES) <= 1 for k in EMPLOYEES for m in TASKS for s in SCENARIOS), name="c3"
        )
        #print("c3")

        # (4) Ensure tasks are performed in ascending order by task number
        m.addConstrs(
            (gp.quicksum(x[(k, r, m+1, s)] for r in CARMOVES) <= gp.quicksum(x[(k,r,m,s)] for r in CARMOVES)
            for k in EMPLOYEES for m in TASKS[:-1] for s in SCENARIOS), name="c4"
        )
        #print("c4")

        # (5) All cars in need of charging must be moved to a charging station within the planning period
        m.addConstrs(
            ((gp.quicksum(x[(k, r, m, s)] for k in EMPLOYEES for r in CHARGING_MOVES_ORIGINATING_IN_NODE[i] for m in TASKS) == INITIAL_NEED_CHARGING[i])
             for i in PARKING_NEED_CHARGING_NODES for s in SCENARIOS), name="c5"
        )
        #print("c5")


        ## Node states

        # (6) Ensure capacity of charging stations is not exceeded in any of the scenarios
        m.addConstrs(
            ((gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in CHARGING_MOVES_ENDING_IN_NODE[i] for m in TASKS) <= INITIAL_AVAILABLE_CAPACITY[i])
             for i in CHARGING_NODES for s in SCENARIOS), name="c6"
        )
        #print("c6")

        # (7) Calculate number of available cars in each parking node by the beginning of the second stage
        m.addConstrs((y[i] == INITIAL_AVAILABLE_CARS[i]
                      + gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in PARKING_MOVES_ENDING_IN_NODE[i] for m in TASKS_FIRST_STAGE)
                    - gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in PARKING_MOVES_ORIGINATING_IN_NODE[i] for m in TASKS_FIRST_STAGE)
                    for i in PARKING_NODES for s in SCENARIOS), name="c7"
        )
        #print("c7")

        ## Demand during planning period

        # (8) number of customer requests served in each parking node in the second stage is less than available cars
        # in the beginning of the second stage
        m.addConstrs((z[(i, s)] <= y[i] + CUSTOMER_DELIVERIES[(i,s)] - gp.quicksum(x[(k,r,m,s)]
                                                      for k in EMPLOYEES
                                                      for r in PARKING_MOVES_ORIGINATING_IN_NODE[i]
                                                      for m in TASKS_SECOND_STAGE)
                      for i in PARKING_NODES for s in SCENARIOS), name="c8")
        #print("c8")


        # (9) number of customer requests served in each parking node in the second stage is less than the number of
        # customer requests in the beginning of the second stage

        m.addConstrs((z[(i, s)] <= CUSTOMER_REQUESTS[(i,s)] for i in PARKING_NODES for s in SCENARIOS), name="c9")
        #print("c9")

        # (10) calculate the number of cars short of the ideal state in every node in each scenario at the end
        # of the planning period
        m.addConstrs((w[(i, s)] >=
                      IDEAL_STATE[i]
                      - INITIAL_AVAILABLE_CARS[i]
                      - gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in PARKING_MOVES_ENDING_IN_NODE[i] for m in TASKS)
                      + gp.quicksum(x[(k,r,m,s)] for k in EMPLOYEES for r in PARKING_MOVES_ORIGINATING_IN_NODE[i] for m in TASKS)
                      + z[(i,s)]
                      - CUSTOMER_DELIVERIES[(i,s)] for i in PARKING_NODES for s in SCENARIOS), name="c10"
                     )
        #print("c10")

        ## Time tracking of node visits


        # (11) ensure that no task can be finished after the end of the planning period
        # m.addConstrs((START_TIME_EMPLOYEE[k]
        #              + gp.quicksum(TRAVEL_TIME[EMPLOYEE_START_LOCATION[k]-1][CARMOVE_ORIGIN[r]-1]*x[(k,r,1,s)] for r in CARMOVES)
        #              + gp.quicksum(RELOCATION_TIME[r] * x[(k, r, m, s)]+(1 / len(CARMOVES))
        #                            * gp.quicksum(TRAVEL_TIME[CARMOVE_DESTINATION[r]-1][CARMOVE_ORIGIN[v]-1]*x[(k,v,m+1,s)] for v in CARMOVES)
        #                            for m in TASKS[:-1] for r in CARMOVES)
        #              + gp.quicksum(RELOCATION_TIME[r]*x[(k,r,len(TASKS),s)] for r in CARMOVES)
        #              <= PLANNING_PERIOD for k in EMPLOYEES for s in SCENARIOS), name="c11")



        m.addConstrs(
                    (t[(k,m,s)]
                     + RELOCATION_TIME[r]*x[(k, r, m, s)]
                     + gp.quicksum(TRAVEL_TIME[CARMOVE_DESTINATION[r]-1][CARMOVE_ORIGIN[v]-1]*x[(k, v, m+1, s)]
                                   for v in CARMOVES)
                     - BIGM[r]*(1-x[(k, r, m, s)])
                     <= t[(k, m+1, s)]
                     for k in EMPLOYEES for r in CARMOVES for m in TASKS[:-1] for s in SCENARIOS), name="c11")

        m.addConstrs(
                    (t[(k,m,s)] <= t[(k,m+1,s)]
                     for k in EMPLOYEES for m in TASKS[:-1] for s in SCENARIOS), name="c12")

        m.addConstrs(((START_TIME_EMPLOYEE[k] + TRAVEL_TIME[EMPLOYEE_START_LOCATION[k] - 1][CARMOVE_ORIGIN[r] - 1]) * x[(k, r, 1, s)]
            <= t[(k, 1, s)] for k in EMPLOYEES for r in CARMOVES for s in SCENARIOS), name="c13")

        m.addConstrs((CARMOVE_START_TIME[r] * x[(k,r,m,s)] <= t[(k, m, s)]
                      for k in EMPLOYEES for r in CARMOVES for m in TASKS for s in SCENARIOS), name="c14")

        m.addConstrs((t[(k, TASKS[-1], s)] + gp.quicksum(RELOCATION_TIME[r]*x[(k, r, TASKS[-1], s)] for r in CARMOVES) <=
                     PLANNING_PERIOD for k in EMPLOYEES for s in SCENARIOS), name="c15")



        #print("Travel time")
        #print(t)
        #t += START_TIME_EMPLOYEE[1] + TRAVEL_TIME[EMPLOYEE_START_LOCATION[1]-1][CARMOVE_ORIGIN[4]-1]
       # print("Start time of task 1: ", t)
        #t += RELOCATION_TIME[4]
        #t += TRAVEL_TIME[CARMOVE_DESTINATION[4] -1][CARMOVE_ORIGIN[2]-1]
        #print("Start time of task 2: ", t)
        #t += RELOCATION_TIME[2]
        #t += TRAVEL_TIME[CARMOVE_DESTINATION[2] - 1][CARMOVE_ORIGIN[9] - 1]
        #print("Start time of task 3: ", t)
        #print(t)


        #x[2, 3, 1, 1] 1(2 --> 10)
        #x[2, 45, 2, 1] 1(6 --> 8)
        #x[2, 1, 3, 1] 1(1 --> 10)


        #x[3, 22, 1, 1] 1(4 --> 11)
        #x[3, 48, 2, 1] 1(7 --> 11)


        #4-->11
        #7-->11

        #print("c11")

        ## Non-anticipativity constraints

        # (15) ensure that all first-stage decisions are equal regardless of scenario
        m.addConstrs((x[(k,r,m,s)] == x[(k,r,m,s+1)]
                      for k in EMPLOYEES
                      for r in CARMOVES
                      for m in TASKS_FIRST_STAGE
                      for s in SCENARIOS[:-1]), name="c15")


        m.addConstrs((t[(k, m, s)] == t[(k, m, s + 1)]
                      for k in EMPLOYEES
                      for m in TASKS_FIRST_STAGE
                      for s in SCENARIOS[:-1]), name="extra")


        return m

    def run_model(m, stochastic=True, reset=False):
        # Optimize Model
        m.write("out.attr")
        try:
            if reset:
                # disable warm start
                m.reset(0)
            m.optimize()
            print("Runtime: ", m.Runtime)

            if m.solCount == 0:
                print("Model is infeasible")
                m.computeIIS()
                m.write("model_iis.ilp")

            ### Print node states ###
            #print("--- INITIAL NODE STATES ---")
            # for i in range(len(PARKING_NODES)):
            #     if stochastic:
            #         print("node: {0}, pstate: {1}, cstate: {2}, istate: {3}, requests: {4}, deliveries: {5}".format(
            #             i+1, cp.pStateP[i],cp.cStateP[i], IDEAL_STATE[i+1], cp.demandP[i], cp.deliveriesP[i]))
            #     else:
            #         print("node: {0}, pstate: {1}, cstate: {2}, istate: {3}, requests: {4:.2f}, deliveries: {5:.2f}".format(
            #             i+1, cp.pStateP[i],cp.cStateP[i], IDEAL_STATE[i+1], round(sum(cp.demandP[i])/len(cp.demandP[i])), round(sum(cp.deliveriesP[i])/len(cp.deliveriesP[i]))))

            #for k,v in EMPLOYEE_START_LOCATION.items():
            #    print("Employee: {}, Node: {}".format(k,v))




            #print("\n--- RESULTS ---")
            x_count, y_count,z_count, w_count = 1,1,1,1
            #for v in m.getVars():
            #    if(v.varName[0] == 't'):
            #       print(v.varName, v.x)
            #         if (v.varName[0] == 'x' and x_count > 0):
            #             print("x[k,r,m,s]: Service employee k performs car-move r as task number m in scenario s")
            #             x_count = 0
            #         if (v.varName[0] == 'y' and y_count > 0):
            #             print("y[i]: Number of cars in node i by the beginning of the second stage")
            #             y_count = 0
            #         elif (v.varName[0] == 'z' and z_count > 0):
            #             print("z[i,s]: Number of customer requests served in second stage in node i in scenario s")
            #             z_count = 0
            #         elif (v.varName[0] == 'w' and w_count > 0):
            #             print("w[i,s]: Number of cars short of ideal state in node i in scenario s")
            #             w_count = 0
            #
            #    if (v.varName[0] == 'x' and v.x > 0):
            #        l = list(map(int, (v.varName[2:-1].split(','))))
            #        print("{0} {1} ({2} --> {3})".format(v.varName, int(v.x),CARMOVE_ORIGIN[l[1]],CARMOVE_DESTINATION[l[1]]))
            #         elif (v.varName[0] != 'x'):
            #             print('%s %g' % (v.varName, v.x))


            print('Obj: %g' % m.objVal)

            #print("--- ROUTES AND SCHEDULES ---")
            #print("Number of tasks in first stage : {}".format(len(TASKS_FIRST_STAGE)))
            #print("Planning period: {}".format(PLANNING_PERIOD))
            employee_routes = {new_list: [] for new_list in EMPLOYEES}
            #print(employee_routes)
            for v in m.getVars():
                if (v.varName[0] == 'x' and v.x > 0):
                    x = list(map(int, (v.varName[2:-1].split(','))))
                    employee_routes[x[0]].append(x[1:])

            # sort by ascending task number, then scenario
            for k in EMPLOYEES:
                employee_routes[k] = sorted(employee_routes[k], key=lambda x: (x[1], x[2]), reverse=False)

            # initialize dataframes of routes and schedules
            df_firststage_routes = pd.DataFrame(columns=["Employee", "Task", "Route", "Travel Time to Task", "Start time", "Relocation Time",
                         "End time"])
            df_secondstage_routes = pd.DataFrame(columns=["Employee", "Task", "Scenario", "Route", "Travel Time to Task", "Start time", "Relocation Time", "End time"])

            # calculate travel times
            for k in EMPLOYEES:
                # employee_routes {k: [r,m,s]}

                # initialize time and scenario dictionaries
                # time for scenario s, t[x[2]]
                t = {s: 0 for s in [x[2] for x in employee_routes[k]]}
                destination_node = {s: 0 for s in [x[2] for x in employee_routes[k]]}
                for x in employee_routes[k]:
                    #First task
                    if x[1] == 1:
                        # Initialize start time of employee
                        t[x[2]] += START_TIME_EMPLOYEE[k]
                        # Bike travel time between start location of employee and origin node of first task
                        tt = TRAVEL_TIME[EMPLOYEE_START_LOCATION[k]-1][CARMOVE_ORIGIN[x[0]]-1]
                    else:
                        # Bike travel time between previous node and current node
                        tt = TRAVEL_TIME[destination_node[x[2]]-1][CARMOVE_ORIGIN[x[0]]-1]

                    # Start time of current task
                    t[x[2]] += tt

                    # Relocation time of car-move
                    rt = RELOCATION_TIME[x[0]]

                    ## Add routes and schedules to dataframe
                    # first stage routes
                    if x[1] <= len(TASKS_FIRST_STAGE):
                        if(x[2] == 1):
                            first_stage_row = [k, x[1], (CARMOVE_ORIGIN[x[0]], CARMOVE_DESTINATION[x[0]]), tt, t[x[2]], rt, t[x[2]] + rt]
                            df_firststage_routes.loc[len(df_firststage_routes)] = first_stage_row

                    # second stage routes
                    else:
                        second_stage_row = [k, x[1], x[2], (CARMOVE_ORIGIN[x[0]], CARMOVE_DESTINATION[x[0]]), tt, t[x[2]], rt, t[x[2]] + rt]
                        df_secondstage_routes.loc[len(df_secondstage_routes)] = second_stage_row
                    #print("Task: {0}, Scenario: {1}, ({2} --> {3}), Travel Time to Task: {4:.1f}, Start time: {5:.1f}, "
                    #      "Relocation Time: {6:.1f}, End time: {7:.1f}".format(
                    #    x[1], x[2], CARMOVE_ORIGIN[x[0]], CARMOVE_DESTINATION[x[0]], tt, t[x[2]], rt, t[x[2]]+rt))

                    # End time
                    t[x[2]] += rt

                    # Update last visited  node in scenario s
                    destination_node[x[2]] = CARMOVE_DESTINATION[x[0]]

            # TODO: sort dataframe with ascending endtime in addition to Employee and, Task and scenario
            pd.set_option('display.width', 320)
            pd.set_option('display.max_columns', 10)
            #print("-------------- First stage routes --------------")
            #print(df_firststage_routes)
            #print("\n-------------- Second stage routes --------------")
            #print(df_secondstage_routes)

            return m, df_firststage_routes, df_secondstage_routes, PLANNING_PERIOD

        except gp.GurobiError as e:
            print('Error code ' + str(e.errno) + ': ' + str(e))

        except AttributeError as e:
            print(e)
            print('Encountered an attribute error')


    stochastic_model = create_model(modelType=0)
    stochastic_model.setParam('TimeLimit', 30*60)
    z_stochastic, firststageroutes, secondstageroutes,  planningperiod = run_model(stochastic_model, stochastic=True, reset=True)
    deterministic_model = create_model(modelType=1)
    z_deterministic = run_model(deterministic_model, stochastic=False)[0]
    #EEV: expectation of expected value. Deterministic first stage values are fixed and then run on the stochastic Model
    eev_model = create_model(modelType=3, input_model=deterministic_model)
    z_eev = run_model(eev_model, stochastic=True)[0]
    vss = z_stochastic.ObjVal - z_eev.ObjVal
    print("Value of Stochastic Solution (VSS)", vss)

    #Write results to file
    resultsfolder = "../Results/4nodes_cold_start/"
    file = filepath.split("/")[-1]
    resultsfile = os.path.join(resultsfolder, file[:-4] + "_results.txt") #str(filepath[-17:-4]) + "_results.txt"

    if not os.path.exists(resultsfolder):
        os.makedirs(resultsfolder)

    f = open(resultsfile, "a+")

    str1 = firststageroutes.to_string(header=True, index=True)
    str2 = secondstageroutes.to_string(header=True, index=True)
    f.write("-------------- First stage routes --------------\n")
    f.write(str1 +"\n")
    f.write("-------------- Second stage routes --------------\n")
    f.write(str2+"\n")

    string = "\nObjVal Stochastic: " + str(z_stochastic.ObjVal) + "\nEEV: " +str(z_eev.ObjVal) + "\nVSS: " +str(vss) + \
             "\nPercentage Improvement: " +str(((z_stochastic.ObjVal/z_eev.ObjVal)-1)*100) +"\nRuntime stochastic: "\
             +str(z_stochastic.Runtime)+"\nPlanning period: "+str(planningperiod)
    f.write(string)

    f.close()

    #Write results to excel file
    firststageroutes = firststageroutes.to_numpy()[:,2]
    secondstageroutes = secondstageroutes.to_numpy()[:,3]
    df_results = pd.DataFrame({
        "Test instance": [file[:-4]],#[filepath[-12:-4]],
        "ObjVal Stochastic": [z_stochastic.ObjVal],
        "ObjBound Stochastic": [z_stochastic.ObjBound],
        "EEV": [z_eev.ObjVal],
        "VSS": [vss],
        "Percentage Improvement": [(z_stochastic.ObjVal/z_eev.ObjVal-1)*100],
        "Runtime Stochastic": [z_stochastic.Runtime],
        "First-stage Routes": [firststageroutes],
        "Second-stage Routes": [secondstageroutes]})

    subdir = directory.split("/")[2]
    excel_filename = '../Results/results_excel/4nodes_cold_start.xlsx'
    append_df_to_excel(excel_filename, df_results, sheet_name=subdir, startrow=None, truncate_sheet=False, index=False)




    #if os.path.exists(excel_filename):
    #    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    #        book = load_workbook(excel_filename)
    #        writer.book = book
    #        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #        df_results.to_excel(writer, sheet_name=subdir, index=False)#, cols=["Test instance","ObjVal Stochastic","ObjBound Stochastic", "EEV", "VSS",  "Percentage Improvement", "Runtime Stochastic"])
    #else:
    #    with pd.ExcelWriter(excel_filename, mode='w') as writer:
    #        df_results.to_excel(writer, sheet_name=subdir, index=False)


    #file = open(filepath, "a")
    #string = "\nObjVal Stochastic: " + str(z_stochastic.ObjVal) + "\nEEV: " +str(z_eev.ObjVal) + "\nVSS: " +str(vss) + \
    #         "\nPercentage Improvement: " +str(((z_stochastic.ObjVal/z_eev.ObjVal)-1)*100) +"\nRuntime stochastic: "+str(z_stochastic.Runtime)
    #file.write(string)
    #file.close()

def main():
    for file in reversed(files):
        print(file)
        run_test_instance(file)

main()
#TODO: investigate why ideal state is not met in all scenarios, shouldnt our system be rigged for worst case scenario demand?