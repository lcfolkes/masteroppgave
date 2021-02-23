import yaml
from openpyxl import load_workbook
import numpy as np
from itertools import product
import pandas as pd

def read_config(config_name: str):
    with open(config_name, 'r') as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)

#cf = read_config('../Gurobi/tests/6nodes/6-3-0-1_a.yaml')

def read_2d_array_to_dict(arr):
    arr = np.array(arr)
    rows = np.arange(1, arr.shape[0]+1) # nodes
    cols = np.arange(1, arr.shape[1]+1) # scenarios
    out_dict = {}
    for (r, c) in product(rows, cols):
        out_dict[(r, c)] = arr[r - 1, c - 1]
    return out_dict

def create_dict_of_indices(indices, item_list):
    d = {}
    for i in indices:
        d[i] = [x+1 for x, e in enumerate(item_list) if e == i]
    return d

def create_car_moves_origin_destination(parking_nodes, charging_nodes, origin_list, destination_list):
    RiPO = {new_list: [] for new_list in parking_nodes}
    RiPD = {new_list: [] for new_list in parking_nodes}
    RiCO = {new_list: [] for new_list in parking_nodes}
    RiCD = {new_list: [] for new_list in charging_nodes}

    for i in range(len(destination_list)):
        # parking moves
        if destination_list[i] in parking_nodes:
            RiPO[origin_list[i]].append(i+1)
            RiPD[destination_list[i]].append(i+1)
        # charging moves
        elif destination_list[i] in charging_nodes:
            RiCO[origin_list[i]].append(i + 1)
            RiCD[destination_list[i]].append(i + 1)

    return RiPO, RiPD, RiCO, RiCD

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        add_header = False
    except FileNotFoundError:
        # file does not exist yet, we will create it
        add_header = True
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, header=add_header, **to_excel_kwargs)

    # save the workbook
    writer.save()
